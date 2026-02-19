#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simulador de Renegociação (Res. CMN 4.966) — versão Python

O que este script faz (espelhando a lógica do simulador em HTML/Excel):
- Lê o arquivo .xlsx e carrega:
  - Aba "page" (operações e campos já calculados/projetados na planilha)
  - Aba "Planilha1" (tabelas de PD PF/PJ em AA2:AC21 e AD2:AF21)
- Busca por CPF/CNPJ e, opcionalmente, inclui o Grupo Econômico
- Calcula alertas (avais, risco>13, dias atraso>30, AD) e o modo "CALCULO MANUAL"
- Calcula cenários:
  - ATIVO: nova provisão = (saldo total ajustado pela entrada) * LGD usada
  - ESTÁGIO 1/2: nova provisão = Σ(EAD_ajustada * LGD * PD) (somente quando CALCULO MANUAL)

Requisitos:
  pip install pandas openpyxl tabulate

Uso rápido:
  python simulador_renegociacao.py --arquivo "Calculadora Renegociação.xlsx" --cpf 12345678901 --tipo PF --grupo

Exemplos:
  # ATIVO com garantia real e entrada
  python simulador_renegociacao.py --arquivo calc.xlsx --cpf 123 --tipo PF --grupo \
    --ativo --garantia --entrada 5000

  # Estágio 1/2 (cálculo manual) com entrada
  python simulador_renegociacao.py --arquivo calc.xlsx --cpf 123 --tipo PJ --grupo \
    --e12 --entrada 10000

Obs.: Este script usa os valores já existentes na aba "page".
Se o seu Excel tiver fórmulas e não tiver sido salvo com valores calculados, abra e salve no Excel antes.
"""

from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from tabulate import tabulate


# ─────────────────────────────────────────────
# Constantes de LGD (centralizadas aqui para
# facilitar atualizações regulatórias futuras)
# ─────────────────────────────────────────────
class LGD:
    ATIVO_GARANTIA_REAL: float = 0.354

    # Estágio 1/2 com garantia real
    E1_GARANTIA_REAL: float = 0.1093
    E2_GARANTIA_REAL: float = 0.1579

    # Sem garantia real (fallback por tipo)
    FALLBACK_PF: float = 0.7662
    FALLBACK_PJ: float = 0.7545


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def only_digits(x: object) -> str:
    s = "" if x is None else str(x)
    return re.sub(r"\D+", "", s)


def clean_str(v: object) -> str:
    """Converte valor para string limpa, tratando NaN/None vindos do pandas (dtype=str)."""
    s = str(v or "").strip()
    return "" if s.lower() in {"nan", "none", "nat"} else s


def to_number(x: object) -> float:
    """
    Converte números vindos do Excel (float/int) e strings pt-BR ou americanas:
    - "1.234,56"  → 1234.56  (pt-BR com separador de milhar)
    - "1234,56"   → 1234.56  (pt-BR sem separador de milhar)
    - "1.5"       → 1.5      (formato americano — NÃO remove o ponto)
    - "35,7%" ou "35.7%" → 0.357
    """
    if x is None:
        return 0.0
    if isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x)):
        return float(x)
    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return 0.0

    pct = False
    if "%" in s:
        pct = True
        s = s.replace("%", "").strip()

    # ── Detectar formato antes de converter (CORRIGIDO) ──────────────────────
    # pt-BR com milhar:  "1.234,56" → tem ponto E vírgula
    # pt-BR sem milhar:  "1234,56"  → tem só vírgula
    # Americano:         "1.5"      → tem só ponto → não mexer
    if "," in s and "." in s:
        # formato pt-BR com separador de milhar: remove ponto, troca vírgula
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        # formato pt-BR sem separador de milhar: só troca vírgula
        s = s.replace(",", ".")
    # else: formato americano ("1.5") — não modifica

    try:
        v = float(s)
    except ValueError:
        return 0.0

    return v / 100.0 if pct else v


def fmt_brl(v: Optional[float]) -> str:
    if v is None:
        return "—"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_pct(v: Optional[float]) -> str:
    if v is None:
        return "—"
    return f"{v * 100:.2f}%".replace(".", ",")


def norm_upper(s: object) -> str:
    return str(s or "").strip().upper()


def validate_cpf_cnpj(cpf_cnpj: str, tipo: str) -> None:
    """Valida se o número de dígitos bate com o tipo PF (11) ou PJ (14)."""
    digits = only_digits(cpf_cnpj)
    expected = 11 if tipo == "PF" else 14
    if len(digits) != expected:
        raise SystemExit(
            f"Atenção: {tipo} espera {expected} dígitos, mas '{cpf_cnpj}' tem {len(digits)}. "
            "Verifique se o tipo (--tipo) está correto."
        )


# ─────────────────────────────────────────────
# Leitura do Excel  (MELHORIA: workbook único)
# ─────────────────────────────────────────────
PAGE_COLS = [
    "CPF/CNPJ",
    "Nome Cliente",
    "Grupo Econômico",
    "Risco CRL",
    "Contrato",
    "Submodalidade Bacen",
    "Operação Renegociada?",
    "Grupo Garantia",
    "Dias em Atraso",
    "Atraso Projetado Final do Mês",
    "Estágio COP",
    "Estágio Projetado",
    "% PD12",
    "% PDVida",
    "% LGD Projetada",
    "Provisão Projetada",
    "Saldo Contábil Bruto Atual",
]


@dataclass
class Operacao:
    cpf: str
    nome: str
    grupo: str
    risco: str
    contrato: str
    submodalidade: str
    renegociada: str
    grupo_garantia: str
    dias_atraso: float
    atraso_proj: float
    estagio_cop: str
    estagio_proj: str
    pd12: float
    pdvida: float
    lgd_proj: float
    prov_proj: float
    saldo: float


def load_excel_data(xlsx_path: str) -> Tuple[
    List[Operacao],
    Dict[str, Dict[str, float]],
    Dict[str, Dict[str, float]],
]:
    """
    Abre o workbook UMA ÚNICA VEZ e carrega:
      - aba 'page'       → lista de Operacao
      - aba 'Planilha1'  → tabelas pd_pf e pd_pj
    (MELHORIA de performance: antes eram duas leituras separadas)
    """
    # ── Validação amigável do arquivo ────────────────────────────────────────
    if not Path(xlsx_path).exists():
        raise SystemExit(f"Arquivo não encontrado: '{xlsx_path}'. Verifique o caminho informado em --arquivo.")

    # ── Ler aba 'page' com pandas ────────────────────────────────────────────
    df = pd.read_excel(xlsx_path, sheet_name="page", dtype=str)
    missing = [c for c in PAGE_COLS if c not in df.columns]
    if missing:
        raise ValueError(
            f'Colunas ausentes na aba "page": {missing}. '
            "Confirme se você está usando o mesmo arquivo do simulador."
        )

    ops: List[Operacao] = []
    for _, row in df.iterrows():
        cpf = only_digits(row.get("CPF/CNPJ"))
        if not cpf:
            continue
        ops.append(
            Operacao(
                cpf=cpf,
                # CORREÇÃO: clean_str() evita "nan" aparecendo como texto ao usuário
                nome=clean_str(row.get("Nome Cliente")),
                grupo=clean_str(row.get("Grupo Econômico")),
                risco=clean_str(row.get("Risco CRL")),
                contrato=clean_str(row.get("Contrato")),
                submodalidade=clean_str(row.get("Submodalidade Bacen")),
                renegociada=clean_str(row.get("Operação Renegociada?")),
                grupo_garantia=clean_str(row.get("Grupo Garantia")),
                dias_atraso=to_number(row.get("Dias em Atraso")),
                atraso_proj=to_number(row.get("Atraso Projetado Final do Mês")),
                estagio_cop=clean_str(row.get("Estágio COP")),
                estagio_proj=clean_str(row.get("Estágio Projetado")),
                pd12=to_number(row.get("% PD12")),
                pdvida=to_number(row.get("% PDVida")),
                lgd_proj=to_number(row.get("% LGD Projetada")),
                prov_proj=to_number(row.get("Provisão Projetada")),
                saldo=to_number(row.get("Saldo Contábil Bruto Atual")),
            )
        )

    # ── Ler tabelas PD com openpyxl (mesmo arquivo, segunda aba) ────────────
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Planilha1" not in wb.sheetnames:
        raise ValueError('Não encontrei a aba "Planilha1".')

    ws = wb["Planilha1"]

    pd_pf: Dict[str, Dict[str, float]] = {}
    pd_pj: Dict[str, Dict[str, float]] = {}

    for r in range(2, 22):
        risco = ws[f"AA{r}"].value
        if risco:
            pd_pf[norm_upper(risco)] = {
                "pd12": to_number(ws[f"AB{r}"].value),
                "pdvida": to_number(ws[f"AC{r}"].value),
            }

    for r in range(2, 22):
        risco = ws[f"AD{r}"].value
        if risco:
            pd_pj[norm_upper(risco)] = {
                "pd12": to_number(ws[f"AE{r}"].value),
                "pdvida": to_number(ws[f"AF{r}"].value),
            }

    wb.close()
    return ops, pd_pf, pd_pj


# ─────────────────────────────────────────────
# Regras de negócio
# ─────────────────────────────────────────────
@dataclass
class Alertas:
    avais: bool
    risco_gt13: bool
    dias_gt30: bool
    ad: bool
    calculo_manual: bool


def compute_alertas(selected_ops: List[Operacao]) -> Alertas:
    avais = any(norm_upper(op.submodalidade) == "AVAIS E FIANÇAS HONRADOS" for op in selected_ops)

    riscos_alerta = {"R13", "R14", "R15", "R16", "R17", "R18", "R19", "R20"}
    risco_gt13 = any(norm_upper(op.risco) in riscos_alerta for op in selected_ops)

    # CORREÇÃO: verifica se ALGUM contrato individualmente tem atraso > 30,
    # não a soma acumulada (que poderia dar falso positivo).
    # ⚠️ Se a regra de negócio for "soma > 30", reverta para:
    #     dias_gt30 = sum(op.dias_atraso for op in selected_ops) > 30
    dias_gt30 = any(op.dias_atraso > 30 for op in selected_ops)

    ad = any(norm_upper(op.submodalidade) == "ADIANTAMENTOS A DEPOSITANTES" for op in selected_ops)

    # calculo_manual é verdadeiro quando NÃO há nenhum dos alertas acima
    calculo_manual = not (risco_gt13 or dias_gt30 or ad)

    return Alertas(avais=avais, risco_gt13=risco_gt13, dias_gt30=dias_gt30, ad=ad, calculo_manual=calculo_manual)


def sum_saldo(selected_ops: List[Operacao]) -> float:
    return sum(op.saldo for op in selected_ops)


def sum_prov_base(selected_ops: List[Operacao]) -> float:
    return sum(op.prov_proj for op in selected_ops)


def pd_lookup(
    risco: str,
    estagio_proj: str,
    tipo: str,
    pd_pf: Dict[str, Dict[str, float]],
    pd_pj: Dict[str, Dict[str, float]],
) -> Optional[float]:
    key = norm_upper(risco)
    table = pd_pf if tipo == "PF" else pd_pj
    t = table.get(key)
    if not t:
        return None
    est = norm_upper(estagio_proj)
    is_e1 = "ESTÁGIO 1" in est or "ESTAGIO 1" in est
    return to_number(t["pd12"] if is_e1 else t["pdvida"])


def lgd_stage12(op: Operacao, garantia_real: bool, tipo: str) -> float:
    est = norm_upper(op.estagio_proj)
    if not garantia_real:
        return op.lgd_proj

    if "ESTÁGIO 1" in est or "ESTAGIO 1" in est:
        return LGD.E1_GARANTIA_REAL
    if "ESTÁGIO 2" in est or "ESTAGIO 2" in est:
        return LGD.E2_GARANTIA_REAL

    # Fallback por tipo (sem estágio identificado)
    return LGD.FALLBACK_PF if tipo == "PF" else LGD.FALLBACK_PJ


def max_lgd_filtrado(selected_ops: List[Operacao]) -> float:
    alvos = {
        "CAPITAL DE GIRO COM PRAZO VENCIMENTO SUPERIOR 365 DIAS",
        "CRÉDITO PESSOAL - SEM CONSIGNAÇÃO EM FOLHA DE PAGAM.",
    }
    m = 0.0
    for op in selected_ops:
        if norm_upper(op.submodalidade) in alvos:
            m = max(m, op.lgd_proj)
    return m


@dataclass
class ResultadoCenario:
    saldo_ajustado: float
    provisao_base: float
    provisao_nova: Optional[float]
    impacto: Optional[float]
    lgd_usada: Optional[float] = None


def calc_ativo(
    selected_ops: List[Operacao],
    flags: Alertas,
    tipo: str,
    entrada: float = 0.0,
    garantia_real: bool = False,
) -> ResultadoCenario:
    saldo_bruto = sum_saldo(selected_ops)
    saldo_ajust = saldo_bruto - entrada
    base = sum_prov_base(selected_ops)

    # CORREÇÃO: usa `not flags.calculo_manual` como single source of truth,
    # evitando duplicar a lógica de alertas aqui.
    deve_calcular = (not flags.calculo_manual) and bool(selected_ops)

    nova: Optional[float] = None
    lgd_usada: Optional[float] = None

    if deve_calcular:
        if garantia_real:
            lgd_usada = LGD.ATIVO_GARANTIA_REAL
        else:
            m = max_lgd_filtrado(selected_ops)
            lgd_usada = max(m, LGD.FALLBACK_PF) if tipo == "PF" else max(m, LGD.FALLBACK_PJ)
        nova = max(0.0, saldo_ajust) * lgd_usada

    impacto = None if nova is None else (nova - base)
    return ResultadoCenario(
        saldo_ajustado=max(0.0, saldo_ajust),
        provisao_base=base,
        provisao_nova=nova,
        impacto=impacto,
        lgd_usada=lgd_usada,
    )


def calc_estagio12(
    selected_ops: List[Operacao],
    flags: Alertas,
    tipo: str,
    pd_pf: Dict[str, Dict[str, float]],
    pd_pj: Dict[str, Dict[str, float]],
    entrada: float = 0.0,
    garantia_real: bool = False,
) -> ResultadoCenario:
    saldo_bruto = sum_saldo(selected_ops)
    saldo_ajust = saldo_bruto - entrada
    saldo_ajust_pos = max(0.0, saldo_ajust)
    fator_entrada = (saldo_ajust_pos / saldo_bruto) if saldo_bruto > 0 else 0.0

    base = sum_prov_base(selected_ops)

    deve_calcular = flags.calculo_manual and bool(selected_ops)

    nova: Optional[float] = None
    if deve_calcular:
        total = 0.0
        for op in selected_ops:
            pd = pd_lookup(op.risco, op.estagio_proj, tipo, pd_pf, pd_pj)
            if pd is None:
                continue
            lgd = lgd_stage12(op, garantia_real, tipo)
            ead = op.saldo * fator_entrada
            total += ead * lgd * pd
        nova = total

    impacto = None if nova is None else (nova - base)
    return ResultadoCenario(
        saldo_ajustado=saldo_ajust_pos,
        provisao_base=base,
        provisao_nova=nova,
        impacto=impacto,
    )


# ─────────────────────────────────────────────
# Seleção por CPF / Grupo
# ─────────────────────────────────────────────
def select_ops(
    ops: List[Operacao],
    cpf_cnpj: str,
    incluir_grupo: bool,
) -> Tuple[List[Operacao], str, int]:
    cpf = only_digits(cpf_cnpj)
    ops_cpf = [o for o in ops if o.cpf == cpf]
    if not ops_cpf:
        raise ValueError("CPF/CNPJ não encontrado na base (aba 'page').")

    grupos = sorted(
        {
            norm_upper(o.grupo)
            for o in ops_cpf
            if norm_upper(o.grupo) and norm_upper(o.grupo) != "NÃO INFORMADO"
        }
    )
    grupo_ativo = grupos[0] if (incluir_grupo and grupos) else ""

    if grupo_ativo:
        # CORREÇÃO: filtra por grupo diretamente (sem concatenar ops_cpf depois),
        # evitando duplicatas quando o CPF já está no grupo.
        ops_grupo = [o for o in ops if norm_upper(o.grupo) == grupo_ativo]
        seen: set = set()
        merged: List[Operacao] = []
        for o in ops_grupo:
            if o.contrato and o.contrato not in {"—", ""}:
                key = f"CT:{o.contrato}"
            else:
                # CORREÇÃO: usa round() para evitar imprecisão de float na chave
                key = f"R:{o.cpf}|{o.submodalidade}|{round(o.saldo, 2)}|{round(o.dias_atraso, 2)}"
            if key in seen:
                continue
            seen.add(key)
            merged.append(o)
        current = merged
    else:
        current = ops_cpf

    tomadores = len({o.cpf for o in current})
    return current, grupo_ativo, tomadores


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────
def main() -> None:
    ap = argparse.ArgumentParser(
        description="Simulador de Renegociação (Res. CMN 4.966) - Python"
    )
    ap.add_argument("--arquivo", required=True, help="Caminho do .xlsx (calculadora)")
    ap.add_argument(
        "--cpf",
        required=True,
        help="CPF/CNPJ do cooperado (somente dígitos ou formatado)",
    )
    ap.add_argument(
        "--tipo",
        choices=["PF", "PJ"],
        default="PF",
        help="Tipo do tomador (PF/PJ)",
    )
    ap.add_argument(
        "--grupo",
        action="store_true",
        help="Incluir operações do Grupo Econômico (quando existir)",
    )
    ap.add_argument(
        "--listar",
        action="store_true",
        help="Lista as operações encontradas (tabela)",
    )
    ap.add_argument(
        "--indices",
        default="",
        help="Selecionar apenas alguns índices (ex: 1,3,5). Se vazio, seleciona tudo.",
    )

    # Cenários
    ap.add_argument("--ativo", action="store_true", help="Calcular cenário ATIVO")
    ap.add_argument(
        "--e12", action="store_true", help="Calcular cenário Estágio 1/2"
    )
    ap.add_argument(
        "--entrada",
        type=float,
        default=0.0,
        help="Valor de entrada (abatimento) para o cenário escolhido",
    )
    ap.add_argument(
        "--garantia",
        action="store_true",
        help="Marcar garantia real no cenário escolhido",
    )

    args = ap.parse_args()

    # MELHORIA: valida CPF/CNPJ x tipo antes de qualquer I/O
    validate_cpf_cnpj(args.cpf, args.tipo)

    # MELHORIA: abre o Excel uma única vez para as duas abas
    ops, pd_pf, pd_pj = load_excel_data(args.arquivo)

    current, grupo_ativo, tomadores = select_ops(ops, args.cpf, args.grupo)

    # Seleção de índices (1-based)
    selected = current
    if args.indices.strip():
        wanted: set = set()
        for part in args.indices.split(","):
            part = part.strip()
            if not part:
                continue
            try:
                wanted.add(int(part))
            except ValueError:
                raise SystemExit(f"Índice inválido em --indices: {part}")
        selected = [op for i, op in enumerate(current, start=1) if i in wanted]

    flags = compute_alertas(selected)

    nome = current[0].nome or "—"
    grupo_exib = grupo_ativo or (
        "" if norm_upper(current[0].grupo) == "NÃO INFORMADO" else (current[0].grupo or "")
    )
    grupo_exib = grupo_exib or "—"

    print("\n=== Identificação ===")
    print(f"CPF/CNPJ: {only_digits(args.cpf)}")
    print(f"Nome:     {nome}")
    print(f"Grupo:    {grupo_exib}")
    print(f"Tipo:     {args.tipo}")
    print(
        f"Operações no conjunto: {len(current)} | "
        f"Selecionadas: {len(selected)} | "
        f"Tomadores no grupo: {tomadores}"
    )

    print("\n=== Alertas ===")
    print(f"Avais:            {'⚠ ALERTA' if flags.avais else '✓ OK'}")
    print(f"Risco > 13:       {'⚠ ALERTA' if flags.risco_gt13 else '✓ OK'}")
    print(f"Dias atraso > 30: {'⚠ ALERTA' if flags.dias_gt30 else '✓ OK'}")
    print(f"AD:               {'⚠ ALERTA' if flags.ad else '✓ OK'}")
    print(f"Modo:             {'CALCULO MANUAL' if flags.calculo_manual else '⚠ ALERTA'}")

    if args.listar:
        df_list = pd.DataFrame(
            [
                {
                    "Sel#": i,
                    "Risco": op.risco,
                    "CPF/CNPJ": op.cpf,
                    "Contrato": op.contrato,
                    "Submodalidade": op.submodalidade,
                    "Renegociada": op.renegociada,
                    "Dias atraso": op.dias_atraso,
                    "Atraso proj.": op.atraso_proj,
                    "Grupo garantia": op.grupo_garantia,
                    "Estágio (COP)": op.estagio_cop,
                    "Estágio proj.": op.estagio_proj,
                    "% PD12": fmt_pct(op.pd12),
                    "% PDVida": fmt_pct(op.pdvida),
                    "% LGD proj.": fmt_pct(op.lgd_proj),
                    "Provisão proj.": fmt_brl(op.prov_proj),
                    "Saldo devedor": fmt_brl(op.saldo),
                }
                for i, op in enumerate(current, start=1)
            ]
        )
        print("\n=== Operações encontradas (aba page) ===")
        print(tabulate(df_list, headers="keys", tablefmt="github", showindex=False))

    if not args.ativo and not args.e12:
        print("\n(Nenhum cenário escolhido. Use --ativo e/ou --e12.)")
        return

    if args.ativo:
        r = calc_ativo(selected, flags, args.tipo, entrada=args.entrada, garantia_real=args.garantia)
        print("\n=== Cenário ATIVO ===")
        print(f"Saldo devedor (ajustado):        {fmt_brl(r.saldo_ajustado)}")
        print(f"Provisão (projetada - base):     {fmt_brl(r.provisao_base)}")
        print(f"Provisão renegociada (calc.):    {fmt_brl(r.provisao_nova)}")
        print(f"Impacto:                         {fmt_brl(r.impacto)}")
        if r.lgd_usada is not None:
            print(f"LGD usada:                       {fmt_pct(r.lgd_usada)}")
        if r.provisao_nova is None:
            print("Obs.: Este cenário só calcula quando houver ALERTA (Risco>13, Dias>30 ou AD).")

    if args.e12:
        r = calc_estagio12(
            selected, flags, args.tipo, pd_pf, pd_pj,
            entrada=args.entrada, garantia_real=args.garantia,
        )
        print("\n=== Cenário Estágio 1/2 ===")
        print(f"Saldo devedor (ajustado):        {fmt_brl(r.saldo_ajustado)}")
        print(f"Provisão (projetada - base):     {fmt_brl(r.provisao_base)}")
        print(f"Provisão renegociada (calc.):    {fmt_brl(r.provisao_nova)}")
        print(f"Impacto:                         {fmt_brl(r.impacto)}")
        if r.provisao_nova is None:
            print("Obs.: Este cenário só calcula em CALCULO MANUAL (sem alertas) e com operações selecionadas.")


if __name__ == "__main__":
    main()
